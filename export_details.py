"""Write all parsed venue details into the details/ folder (Excel + CSV)."""

from __future__ import annotations

import csv
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ingest import DAYS, PERIOD_TIMES, load_all, occupancy_rows, year_sort

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(ROOT, "data")
DETAILS_DIR = os.path.join(ROOT, "details")
CONFIG_PATH = os.path.join(ROOT, "config.json")
PUBLIC_LINK = "https://srm-venue-dashboard-bsvhpqvfgqsftcnphjhphd.streamlit.app/"

title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1A237E")
year_font = Font(name="Calibri", size=12, bold=True, color="0D47A1")
year_fill = PatternFill("solid", fgColor="E3F2FD")
header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="37474F")
venue_font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
rot_font = Font(name="Calibri", size=9, color="E65100")
normal_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
thin = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
DAY_FULL = {"MON": "MONDAY", "TUE": "TUESDAY", "WED": "WEDNESDAY", "THU": "THURSDAY", "FRI": "FRIDAY"}
DAY_FILLS = {
    "MON": PatternFill("solid", fgColor="FFF3E0"),
    "TUE": PatternFill("solid", fgColor="E8F5E9"),
    "WED": PatternFill("solid", fgColor="E3F2FD"),
    "THU": PatternFill("solid", fgColor="FCE4EC"),
    "FRI": PatternFill("solid", fgColor="F3E5F5"),
}


def safe_sheet(name: str) -> str:
    cleaned = re.sub(r"[\[\]\*\/\\\?\:]", "-", name)[:31]
    return cleaned or "Sheet"


def write_dept_sheet(wb, sheet_name, full_name, records, semester):
    ws = wb.create_sheet(safe_sheet(sheet_name))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    c = ws.cell(1, 1, f"SRM Ramapuram — {full_name} — {semester}")
    c.font, c.fill, c.alignment = title_font, title_fill, center
    if not records:
        ws.cell(3, 1, "No data")
        return
    row = 3
    for yr in sorted({r["year"] for r in records}, key=year_sort):
        yr_recs = [r for r in records if r["year"] == yr]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws.cell(row, 1, f"{yr} ({len(yr_recs)} sections)")
        c.font, c.fill, c.alignment, c.border = year_font, year_fill, left, thin
        row += 1
        for day in DAYS:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            c = ws.cell(row, 1, DAY_FULL[day])
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="546E7A")
            c.alignment, c.border = center, thin
            row += 1
            headers = ["Section", "Venue"] + [f"P{p}\n{PERIOD_TIMES[p]}" for p in range(1, 9)] + ["Rotation"]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row, ci, h)
                cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, thin
            row += 1
            fill = DAY_FILLS[day]
            for rec in yr_recs:
                tt = rec.get("timetable", {}).get(day, {})
                vals = [
                    rec["section"],
                    rec["venue"],
                    *[tt.get(p, "") or "-" for p in range(1, 9)],
                    ", ".join(rec.get("rotation_venues") or []) or "-",
                ]
                fonts = [bold_font, venue_font] + [normal_font] * 8 + [rot_font]
                aligns = [left, left] + [center] * 8 + [left]
                for ci, (val, font, al) in enumerate(zip(vals, fonts, aligns), 1):
                    cell = ws.cell(row, ci, val)
                    cell.font, cell.fill, cell.alignment, cell.border = font, fill, al, thin
                row += 1
        row += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["K"].width = 36


def write_summary(ws, rows, semester):
    ws.title = "Summary"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c = ws.cell(1, 1, f"SRM Ramapuram — All Programme Venue Summary ({semester})")
    c.font, c.fill, c.alignment = title_font, title_fill, center
    headers = ["S.No", "Department", "Year", "Section", "Venue / Room", "Faculty Advisor", "Rotation Venues"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(3, ci, h)
        cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, thin
    for i, rec in enumerate(rows, 1):
        vals = [i, rec["dept"], rec["year"], rec["section"], rec["venue"], rec.get("fa", ""), rec["rotation"]]
        fonts = [normal_font, bold_font, normal_font, normal_font, venue_font, normal_font, rot_font]
        for ci, (val, font) in enumerate(zip(vals, fonts), 1):
            cell = ws.cell(3 + i, ci, val)
            cell.font, cell.alignment, cell.border = font, left if ci > 1 else center, thin
    widths = [6, 16, 10, 28, 32, 36, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_csv(path, records, dept_id):
    fieldnames = ["programme", "year", "section", "venue", "faculty", "day", "period", "time", "subject", "rotation_venues"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for rec in records:
            rot = "; ".join(rec.get("rotation_venues") or [])
            for day in DAYS:
                tt = rec.get("timetable", {}).get(day, {})
                for p in range(1, 9):
                    w.writerow(
                        {
                            "programme": dept_id,
                            "year": rec.get("year", ""),
                            "section": rec.get("section", ""),
                            "venue": rec.get("venue", ""),
                            "faculty": rec.get("fa", ""),
                            "day": day,
                            "period": p,
                            "time": PERIOD_TIMES[p],
                            "subject": tt.get(p, ""),
                            "rotation_venues": rot,
                        }
                    )


def main():
    os.makedirs(DETAILS_DIR, exist_ok=True)
    cfg = json.load(open(CONFIG_PATH, encoding="utf-8"))
    semester = cfg.get("semester", "2026-27 Odd Semester")
    catalog, errors = load_all(DATA_DIR, cfg["departments"])
    meta = {d["id"]: d for d in cfg["departments"]}

    summary_rows = []
    wb = Workbook()
    for d in cfg["departments"]:
        recs = catalog.get(d["id"], [])
        write_dept_sheet(wb, d["id"], d["name"], recs, semester)
        write_csv(os.path.join(DETAILS_DIR, f"{d['id'].replace(' ', '_').replace('&', 'and')}.csv"), recs, d["id"])
        for rec in recs:
            summary_rows.append(
                {
                    "dept": d["id"],
                    "year": rec.get("year", ""),
                    "section": rec.get("section", ""),
                    "venue": rec.get("venue", ""),
                    "fa": rec.get("fa", ""),
                    "rotation": ", ".join(rec.get("rotation_venues") or []) or "-",
                }
            )
        print(f"{d['id']:12} {len(recs):3} sections")

    write_summary(wb.active, summary_rows, semester)
    xlsx = os.path.join(DETAILS_DIR, "All_Venue_Details.xlsx")
    wb.save(xlsx)

    occ = occupancy_rows(catalog)
    occ_path = os.path.join(DETAILS_DIR, "occupancy_by_period.csv")
    with open(occ_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["dept", "year", "section", "venue", "home_venue", "day", "period", "subject", "fa"])
        w.writeheader()
        w.writerows(occ)

    n_sec = sum(len(v) for v in catalog.values())
    note = os.path.join(DETAILS_DIR, "FOLDER.txt")
    with open(note, "w", encoding="utf-8") as f:
        f.write(
            f"SRM Ramapuram venue details ({semester})\n"
            "==================================================\n\n"
            f"Public link (anyone can view, no login):\n{PUBLIC_LINK}\n\n"
            f"Programmes: {len(cfg['departments'])}\n"
            f"Sections:   {n_sec}\n"
            f"Period rows:{len(occ)}\n\n"
            "This folder:\n"
            "  All_Venue_Details.xlsx  — one sheet per programme + Summary\n"
            "  occupancy_by_period.csv — who is in which room each period\n"
            "  <programme>.csv         — day/period rows for that programme\n\n"
            "Source timetables live in ../data/\n"
            "App code lives in ../app.py and ../ingest.py\n"
        )
        if errors:
            f.write("\nParser notes:\n")
            for k, msg in errors.items():
                f.write(f"  {k}: {msg}\n")

    print("Wrote", xlsx)
    print("Wrote", occ_path)
    print("Wrote", note)
    if errors:
        print("ERRORS", errors)


if __name__ == "__main__":
    main()
